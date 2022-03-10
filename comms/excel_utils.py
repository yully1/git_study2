"""
----------------------------------------------------
@Time  :  2022/2/25 19:33
@Author  :  shina
@File  :  excel_test03.PY
----------------------------------------------------
"""
import openpyxl  # 用来操作excel

"""
将测试数据封装成对象

"""


class CaseData:  # 将dict_case字典遍历成对象，获取对象的属性值
    def __init__(self, dict_case):
        for i in dict_case.items():
            setattr(self, i[0], i[1])


class ReadExcel:  # 封装excel读取所有方法
    @classmethod
    def read_data_all(cls, file_name, sheet_name):  # 读取所有内容
        workbook = openpyxl.load_workbook(file_name)  # 读取工作簿
        sheet = workbook[sheet_name]  # 获取操作指定的sheet页
        rows = list(sheet.rows)  # 获取所有单元格
        all_case = []  # 定义一个空列表

        # 读取表头数据
        titles = []
        for cell in rows[0]:  # rows[0]:代表第一行数据
            titles.append(cell.value)  # 将第一行单元格的值增加到titles
        # print(titles)#['case_id','case_title','interface','url'......]
        # 遍历其他行数据，和表头打包转换成字典，反射到对象的实例属性，并且把对象存到all_case列表里面
        for row in rows[1:]:  # row代表除了表头的每一行数据
            data = []
            for v in row:
                data.append(v.value)
            # print(data)[1,'正常流程','user_login']
            case_data = dict(zip(titles, data))
            cd = CaseData(case_data)
            all_case.append(cd)
        return all_case

    # 读取指定的行
    @classmethod
    def read_data_pl(cls, file_name, sheet_name, begin_row, end_row):  # 读取指定行
        workbook = openpyxl.load_workbook(file_name)  # 读取工作簿
        sheet = workbook[sheet_name]  # 获取操作指定的sheet页
        rows = list(sheet.rows)  # 获取所有单元格
        all_case = []  # 定义一个空列表

        # 读取表头数据
        titles = []
        for cell in rows[0]:  # rows[0]:代表第一行数据
            titles.append(cell.value)  # 将第一行单元格的值增加到titles
        # print(titles)#['case_id','case_title','interface','url'......]
        # 遍历其他行数据，和表头打包转换成字典，反射到对象的实例属性，并且把对象存到all_case列表里面
        for row in rows[begin_row:end_row + 1]:  # row代表除了表头指定的行数据
            data = []
            for v in row:
                data.append(v.value)
            case_data = dict(zip(titles, data))
            cd = CaseData(case_data)
            all_case.append(cd)
        return all_case

    # 写数据封装
    @classmethod
    def write_data(cls, file_name, sheet_name, row, column, value):
        """

        :param file_name: 需要被写入的文件
        :param sheet_name: 需要被写入的sheet页
        :param row: 需要被写入的行
        :param column: 需要被写入的列
        :param value: 需要被写入的值
        """
        wb = openpyxl.load_workbook(file_name)
        sh = wb[sheet_name]
        sh.cell(row=row + 1, column=column, value=value)
        wb.save(file_name)  # 保存工作


